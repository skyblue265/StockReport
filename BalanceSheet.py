import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook


def CreateExcelFile(path, fileName):
    wb = Workbook()
    ws = wb.active
    ws.title = fileName

    wb.save( fileName = f"{path}\\{fileName}.xlsx")
    return f"{path}\\{fileName}.xlsx"

# def WriteToExcel(excelPath, workSheetName,x, y , value):
#     print(f"{x}, {y} : {value}")

def WriteToExcel(excelPath, worksheet, x, y, value):
    workbook = openpyxl.load_workbook(excelPath)
    worksheet = workbook[worksheet]
    worksheet[f"{x}{y}"] = value
    workbook.save(excelPath)
#
#excelPath = 'C:\\Users\\user\\Desktop\\data.xlsx'
#workSheetName = ' 1 '
#

def CheckStockExist(stockNo):
    cookies = {
        'jcsession': 'jHttpSession @7dcc312a',
        '_ga': 'GA1.3.1282828403.1602870641',
        '_gid': 'GA1.3.2431559586.1604496102',
        'newmops2': 'co_id % 3D2448 % 7Cyear % 3D % 7Cseason % 3D % 7C',
        '_gat': '1',
    }

    headers = {
        'Connection': 'keep-alive',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.243 Safari/537.36',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Accept': '*/*',
        'Origin': 'https://mops.twse.com.tw',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Dest': 'empty',
        'Referer': 'https://mops.twse.com.tw/mops/web/t05st31',
        'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7',
    }

    data = {
      'encodeURIComponent': '1 ',
      'step': '1 ',
      'firstin': '1 ',
      'off': '1 ',
      'keyword4': ' ',
      'code1': ' ',
      'TYPEK2': ' ',
      'checkbtn': ' ',
      'queryName': 'co_id ',
      'inpuType': 'co_id ',
      'TYPEK': 'all ',
      'isnew': 'true ',
      'co_id': stockNo,
      'year': ' ',
      'season': ''
    }

    response = requests.post('https://mops.twse.com.tw/mops/web/ajax_t05st31', headers=headers, cookies=cookies, data=data)
    response.encoding = 'big-5'
    if '不繼續公開發行' not in response.text:
        return True
    else:
        return False
    #print(response.text)


def DownloadData(stockNo, excelPath, workSheetName):
    cookies = {
        'jcsession': 'jHttpSession @7dcc312a',
        '_ga': 'GA1.3.1282828403.1602870641',
        '_gid': 'GA1.3.2431559586.1604496102',
        'newmops2': 'co_id % 3D2448 % 7Cyear % 3D % 7Cseason % 3D % 7C',
        '_gat': '1',
    }

    headers = {
        'Connection': 'keep-alive',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.243 Safari/537.36',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Accept': '*/*',
        'Origin': 'https://mops.twse.com.tw',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Dest': 'empty',
        'Referer': 'https://mops.twse.com.tw/mops/web/t05st31',
        'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7',
    }

    data = {
        'encodeURIComponent': '1 ',
        'step': '1 ',
        'firstin': '1 ',
        'off': '1 ',
        'keyword4': ' ',
        'code1': ' ',
        'TYPEK2': ' ',
        'checkbtn': ' ',
        'queryName': 'co_id ',
        'inpuType': 'co_id ',
        'TYPEK': 'all ',
        'isnew': 'true ',
        'co_id': stockNo,
        'year': ' ',
        'season': ''
    }

    response = requests.post('https://mops.twse.com.tw/mops/web/ajax_t05st31', headers=headers, cookies=cookies,
                             data=data)
    response.encoding = 'big-5'
    soup = BeautifulSoup(response.text, 'html.parser')

    alltblHead = soup.find_all("td", class_="tblHead")
    for item in alltblHead:
        #print(item.text.strip())
        WriteToExcel(excelPath, workSheetName,'E', 2, item.text.strip())

    allTD= soup.find_all('b')
    for item in range(len(allTD)):
        if item ==4:
            WriteToExcel(excelPath, workSheetName,'C', 1, allTD[item].text.strip())
        if item ==5:
            WriteToExcel(excelPath, workSheetName,'A', 3, allTD[item].text.strip())
        if item ==6:
            WriteToExcel(excelPath, workSheetName,'B', 3, allTD[item].text.strip())
        if item ==7:
            WriteToExcel(excelPath, workSheetName,'D', 3, allTD[item].text.strip())
        if item ==8:
            WriteToExcel(excelPath, workSheetName,'B', 4, allTD[item].text.strip())
        if item ==9:
            WriteToExcel(excelPath, workSheetName,'C', 4, allTD[item].text.strip())
        if item ==10:
            WriteToExcel(excelPath, workSheetName,'D', 4, allTD[item].text.strip())
        if item ==11:
            WriteToExcel(excelPath, workSheetName,'E', 4, allTD[item].text.strip())

    allTD = soup.find_all('td')

    for item in range(len(allTD)):
        #print(allTD[item].text.strip())

        if allTD[item].text.strip() =='資產':
            #print (item)
            WriteToExcel(excelPath, workSheetName,'A', 5, '        資產')
            WriteToExcel(excelPath, workSheetName,'B', 5, allTD[item+1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 5, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 5, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 5, allTD[item + 4].text.strip())

        if allTD[item].text.strip() =='流動資產'and item == 9:
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 6, '        流動資產')
            WriteToExcel(excelPath, workSheetName,'B', 6, allTD[item+1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 6, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 6, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 6, allTD[item + 4].text.strip())

        if allTD[item].text.strip() =='現金及約當現金':
            WriteToExcel(excelPath, workSheetName,'A', 7, '          現金及約當現金')
            WriteToExcel(excelPath, workSheetName,'B', 7, allTD[item+1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 7, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 7, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 7, allTD[item + 4].text.strip())

        if allTD[item].text.strip() =='公平價值變動列入損益之金融資產-流動':
            WriteToExcel(excelPath, workSheetName,'A', 8, '          公平價值變動列入損益之金融資產-流動')
            WriteToExcel(excelPath, workSheetName,'B', 8, allTD[item+1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 8, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 8, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 8, allTD[item + 4].text.strip())

        if allTD[item].text.strip() =='應收票據淨額':
            WriteToExcel(excelPath, workSheetName,'A', 9, '          應收票據淨額')
            WriteToExcel(excelPath, workSheetName,'B', 9, allTD[item+1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 9, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 9, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 9, allTD[item + 4].text.strip())

        if allTD[item].text.strip() =='應收帳款淨額':
            WriteToExcel(excelPath, workSheetName,'A', 10, '          應收帳款淨額')
            WriteToExcel(excelPath, workSheetName,'B', 10, allTD[item+1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 10, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 10, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 10, allTD[item + 4].text.strip())

        if allTD[item].text.strip() =='應收帳款–關係人淨額':
            WriteToExcel(excelPath, workSheetName,'A', 11, '          應收帳款–關係人淨額')
            WriteToExcel(excelPath, workSheetName,'B', 11, allTD[item+1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 11, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 11, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 11, allTD[item + 4].text.strip())

        if allTD[item].text.strip() =='其他應收款':
            WriteToExcel(excelPath, workSheetName,'A', 12, '          其他應收款')
            WriteToExcel(excelPath, workSheetName,'B', 12, allTD[item+1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 12, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 12, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 12, allTD[item + 4].text.strip())

        if allTD[item].text.strip() =='其他應收款–關係人':
            WriteToExcel(excelPath, workSheetName,'A', 13, '          其他應收款–關係人')
            WriteToExcel(excelPath, workSheetName,'B', 13, allTD[item+1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 13, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 13, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 13, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '其他金融資產-流動':
            WriteToExcel(excelPath, workSheetName,'A', 14, '          其他金融資產-流動')
            WriteToExcel(excelPath, workSheetName,'B', 14, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 14, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 14, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 14, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '存 貨':
            WriteToExcel(excelPath, workSheetName,'A', 15, '          存 貨')
            WriteToExcel(excelPath, workSheetName,'B', 15, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 15, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 15, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 15, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '預付款項':
            WriteToExcel(excelPath, workSheetName,'A', 16, '          預付款項')
            WriteToExcel(excelPath, workSheetName,'B', 16, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 16, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 16, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 16, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '其他流動資產':
            WriteToExcel(excelPath, workSheetName,'A', 17, '          其他流動資產')
            WriteToExcel(excelPath, workSheetName,'B', 17, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 17, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 17, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 17, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '流動資產'and item == 69:
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 18, '            流動資產')
            WriteToExcel(excelPath, workSheetName,'B', 18, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 18, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 18, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 18, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '基金及投資'and item == 74:
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 19, '        基金及投資')
            WriteToExcel(excelPath, workSheetName,'B', 19, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 19, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 19, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 19, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '備供出售金融資產-非流動':
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 20, '          備供出售金融資產-非流動')
            WriteToExcel(excelPath, workSheetName,'B', 20, allTD[item+1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 20, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 20, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 20, allTD[item + 4].text.strip())


        if allTD[item].text.strip() == '以成本衡量之金融資產-非流動':
            WriteToExcel(excelPath, workSheetName,'A', 21, '          以成本衡量之金融資產-非流動')
            WriteToExcel(excelPath, workSheetName,'B', 21, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 21, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 21, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 21, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '採權益法之長期股權投資':
            WriteToExcel(excelPath, workSheetName,'A', 22, '            採權益法之長期股權投資')
            WriteToExcel(excelPath, workSheetName,'B', 22, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 22, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 22, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 22, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '投資':
            WriteToExcel(excelPath, workSheetName,'A', 23, '          投資')
            WriteToExcel(excelPath, workSheetName,'B', 23, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 23, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 23, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 23, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '基金及投資' and item == 99:
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 24, '            基金及投資')
            WriteToExcel(excelPath, workSheetName,'B', 24, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 24, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 24, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 24, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '固定資產' and item == 104:
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 25, '        固定資產')
            WriteToExcel(excelPath, workSheetName,'B', 25, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 25, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 25, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 25, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '成本' and item == 109 :
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 26, '        成本')
            WriteToExcel(excelPath, workSheetName,'B', 26, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 26, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 26, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 26, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '房屋及建築':
            WriteToExcel(excelPath, workSheetName,'A', 27, '          房屋及建築')
            WriteToExcel(excelPath, workSheetName,'B', 27, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 27, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 27, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 27, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '機器設備':
            WriteToExcel(excelPath, workSheetName,'A', 28, '          機器設備')
            WriteToExcel(excelPath, workSheetName,'B', 28, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 28, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 28, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 28, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '運輸設備':
            WriteToExcel(excelPath, workSheetName,'A', 29, '          運輸設備')
            WriteToExcel(excelPath, workSheetName,'B', 29, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 29, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 29, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 29, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '辦公設備':
            WriteToExcel(excelPath, workSheetName,'A', 30, '          辦公設備')
            WriteToExcel(excelPath, workSheetName,'B', 30, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 30, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 30, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 30, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '租賃改良':
            WriteToExcel(excelPath, workSheetName,'A', 31, '          租賃改良')
            WriteToExcel(excelPath, workSheetName,'B', 31, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 31, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 31, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 31, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '固定資產成本合計':
            WriteToExcel(excelPath, workSheetName,'A', 32, '          固定資產成本合計')
            WriteToExcel(excelPath, workSheetName,'B', 32, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 32, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 32, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 32, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '累積折舊':
            WriteToExcel(excelPath, workSheetName,'A', 33, '          累積折舊')
            WriteToExcel(excelPath, workSheetName,'B', 33, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 33, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 33, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 33, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '累計減損-固定資產':
            WriteToExcel(excelPath, workSheetName,'A', 34, '          累計減損-固定資產')
            WriteToExcel(excelPath, workSheetName,'B', 34, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 34, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 34, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 34, allTD[item + 4].text.strip())


        if allTD[item].text.strip() == '未完工程及預付設備款':
            WriteToExcel(excelPath, workSheetName,'A', 35, '          未完工程及預付設備款')
            WriteToExcel(excelPath, workSheetName,'B', 35, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 35, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 35, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 35, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '固定資產淨額':
            WriteToExcel(excelPath, workSheetName,'A', 36, '            固定資產淨額')
            WriteToExcel(excelPath, workSheetName,'B', 36, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 36, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 36, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 36, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '無形資產':
            WriteToExcel(excelPath, workSheetName,'A', 37, '        無形資產')
            WriteToExcel(excelPath, workSheetName,'B', 37, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 37, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 37, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 37, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '專 利 權':
            WriteToExcel(excelPath, workSheetName,'A', 38, '          專 利 權')
            WriteToExcel(excelPath, workSheetName,'B', 38, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 38, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 38, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 38, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '遞延退休金成本':
            WriteToExcel(excelPath, workSheetName,'A', 39, '          遞延退休金成本')
            WriteToExcel(excelPath, workSheetName,'B', 39, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 39, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 39, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 39, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '其他無形資產':
            WriteToExcel(excelPath, workSheetName,'A', 40, '          其他無形資產')
            WriteToExcel(excelPath, workSheetName,'B', 40, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 40, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 40, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 40, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '無形資產合計':
            WriteToExcel(excelPath, workSheetName,'A', 41, '            無形資產合計')
            WriteToExcel(excelPath, workSheetName,'B', 41, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 41, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 41, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 41, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '其他資產':
            WriteToExcel(excelPath, workSheetName,'A', 42, '        其他資產')
            WriteToExcel(excelPath, workSheetName,'B', 42, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 42, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 42, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 42, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '存出保證金':
            WriteToExcel(excelPath, workSheetName,'A', 43, '          存出保證金')
            WriteToExcel(excelPath, workSheetName,'B', 43, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 43, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 43, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 43, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '遞延費用':
            WriteToExcel(excelPath, workSheetName,'A', 44, '          遞延費用')
            WriteToExcel(excelPath, workSheetName,'B', 44, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 44, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 44, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 44, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '遞延所得稅資產-非流動':
            WriteToExcel(excelPath, workSheetName,'A', 45, '          遞延所得稅資產-非流動')
            WriteToExcel(excelPath, workSheetName,'B', 45, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 45, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 45, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 45, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '其他資產合計':
            WriteToExcel(excelPath, workSheetName,'A', 46, '            其他資產合計')
            WriteToExcel(excelPath, workSheetName,'B', 46, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 46, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 46, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 46, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '資產總計':
            WriteToExcel(excelPath, workSheetName,'A', 47, '            資產總計')
            WriteToExcel(excelPath, workSheetName,'B', 47, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 47, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 47, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 47, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '負債及股東權益':
            WriteToExcel(excelPath, workSheetName,'A', 48, '        負債及股東權益')
            WriteToExcel(excelPath, workSheetName,'B', 48, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 48, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 48, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 48, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '負債':
            WriteToExcel(excelPath, workSheetName,'A', 49, '        負債')
            WriteToExcel(excelPath, workSheetName,'B', 49, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 49, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 49, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 49, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '流動負債' and item == 229:
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 50, '        流動負債')
            WriteToExcel(excelPath, workSheetName,'B', 50, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 50, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 50, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 50, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '短期借款':
            WriteToExcel(excelPath, workSheetName,'A', 51, '          短期借款')
            WriteToExcel(excelPath, workSheetName,'B', 51, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 51, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 51, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 51, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '公平價值變動列入損益之金融負債-流動':
            WriteToExcel(excelPath, workSheetName,'A', 52, '          公平價值變動列入損益之金融負債-流動')
            WriteToExcel(excelPath, workSheetName,'B', 52, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 52, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 52, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 52, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '應付票據':
            WriteToExcel(excelPath, workSheetName,'A', 53, '          應付票據')
            WriteToExcel(excelPath, workSheetName,'B', 53, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 53, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 53, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 53, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '應付帳款':
            WriteToExcel(excelPath, workSheetName,'A', 54, '          應付帳款')
            WriteToExcel(excelPath, workSheetName,'B', 54, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 54, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 54, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 54, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '應付帳款–關係人':
            WriteToExcel(excelPath, workSheetName,'A', 55, '          應付帳款–關係人')
            WriteToExcel(excelPath, workSheetName,'B', 55, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 55, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 55, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 55, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '應付所得稅':
            WriteToExcel(excelPath, workSheetName,'A', 56, '          應付所得稅')
            WriteToExcel(excelPath, workSheetName,'B', 56, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 56, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 56, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 56, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '應付費用':
            WriteToExcel(excelPath, workSheetName,'A', 57, '          應付費用')
            WriteToExcel(excelPath, workSheetName,'B', 57, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 57, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 57, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 57, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '其他應付款項':
            WriteToExcel(excelPath, workSheetName,'A', 58, '          其他應付款項')
            WriteToExcel(excelPath, workSheetName,'B', 58, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 58, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 58, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 58, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '預收款項':
            WriteToExcel(excelPath, workSheetName,'A', 59, '          預收款項')
            WriteToExcel(excelPath, workSheetName,'B', 59, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 59, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 59, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 59, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '一年或一營業週期內到期長期負債':
            WriteToExcel(excelPath, workSheetName,'A', 60, '          一年或一營業週期內到期長期負債')
            WriteToExcel(excelPath, workSheetName,'B', 60, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 60, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 60, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 60, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '其他流動負債':
            WriteToExcel(excelPath, workSheetName,'A', 61, '          其他流動負債')
            WriteToExcel(excelPath, workSheetName,'B', 61, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 61, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 61, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 61, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '流動負債' and item == 289:
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 62, '            流動負債')
            WriteToExcel(excelPath, workSheetName,'B', 62, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 62, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 62, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 62, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '長期負債' and item == 294:
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 63, '        長期負債')
            WriteToExcel(excelPath, workSheetName,'B', 63, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 63, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 63, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 63, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '應付公司債':
            WriteToExcel(excelPath, workSheetName,'A', 64, '          應付公司債')
            WriteToExcel(excelPath, workSheetName,'B', 64, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 64, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 64, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 64, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '長期遞延收入':
            WriteToExcel(excelPath, workSheetName,'A', 65, '          長期遞延收入')
            WriteToExcel(excelPath, workSheetName,'B', 65, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 65, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 65, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 65, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '長期負債' and item == 309:
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 66, '            長期負債')
            WriteToExcel(excelPath, workSheetName,'B', 66, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 66, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 66, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 66, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '各項準備':
            WriteToExcel(excelPath, workSheetName,'A', 67, '        各項準備')
            WriteToExcel(excelPath, workSheetName,'B', 67, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 67, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 67, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 67, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '其他負債':
            WriteToExcel(excelPath, workSheetName,'A', 68, '        其他負債')
            WriteToExcel(excelPath, workSheetName,'B', 68, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 68, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 68, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 68, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '退休金準備／應計退休金負債':
            WriteToExcel(excelPath, workSheetName,'A', 69, '          退休金準備／應計退休金負債')
            WriteToExcel(excelPath, workSheetName,'B', 69, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 69, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 69, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 69, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '存入保證金':
            WriteToExcel(excelPath, workSheetName,'A', 70, '          存入保證金')
            WriteToExcel(excelPath, workSheetName,'B', 70, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 70, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 70, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 70, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '其他負債–其他':
            WriteToExcel(excelPath, workSheetName,'A', 71, '          其他負債–其他')
            WriteToExcel(excelPath, workSheetName,'B', 71, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 71, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 71, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 71, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '其他負債合計':
            WriteToExcel(excelPath, workSheetName,'A', 72, '            其他負債合計')
            WriteToExcel(excelPath, workSheetName,'B', 72, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 72, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 72, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 72, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '負債總計':
            WriteToExcel(excelPath, workSheetName,'A', 73, '            負債總計')
            WriteToExcel(excelPath, workSheetName,'B', 73, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 73, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 73, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 73, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '股東權益':
            WriteToExcel(excelPath, workSheetName,'A', 74, '        股東權益')
            WriteToExcel(excelPath, workSheetName,'B', 74, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 74, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 74, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 74, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '股本':
            WriteToExcel(excelPath, workSheetName,'A', 75, '        股本')
            WriteToExcel(excelPath, workSheetName,'B', 75, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 75, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 75, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 75, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '普通股股本':
            WriteToExcel(excelPath, workSheetName,'A', 76, '          普通股股本')
            WriteToExcel(excelPath, workSheetName,'B', 76, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 76, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 76, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 76, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '預收股本':
            WriteToExcel(excelPath, workSheetName,'A', 77, '          預收股本')
            WriteToExcel(excelPath, workSheetName,'B', 77, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 77, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 77, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 77, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '資本公積':
            WriteToExcel(excelPath, workSheetName,'A', 78, '        資本公積')
            WriteToExcel(excelPath, workSheetName,'B', 78, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 78, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 78, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 78, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '資本公積–發行溢價':
            WriteToExcel(excelPath, workSheetName,'A', 79, '          資本公積–發行溢價')
            WriteToExcel(excelPath, workSheetName,'B', 79, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 79, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 79, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 79, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '資本公積–庫藏股票交易':
            WriteToExcel(excelPath, workSheetName,'A', 80, '          資本公積–庫藏股票交易')
            WriteToExcel(excelPath, workSheetName,'B', 80, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 80, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 80, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 80, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '資本公積–長期投資':
            WriteToExcel(excelPath, workSheetName,'A', 81, '          資本公積–長期投資')
            WriteToExcel(excelPath, workSheetName,'B', 81, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 81, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 81, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 81, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '資本公積-員工認股權':
            WriteToExcel(excelPath, workSheetName,'A', 82, '          資本公積-員工認股權')
            WriteToExcel(excelPath, workSheetName,'B', 82, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 82, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 82, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 82, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '資本公積-認股權':
            WriteToExcel(excelPath, workSheetName,'A', 83, '          資本公積-認股權')
            WriteToExcel(excelPath, workSheetName,'B', 83, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 83, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 83, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 83, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '資本公積–其 他':
            WriteToExcel(excelPath, workSheetName,'A', 84, '          資本公積–其 他')
            WriteToExcel(excelPath, workSheetName,'B', 84, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 84, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 84, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 84, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '資本公積合計':
            WriteToExcel(excelPath, workSheetName,'A', 85, '            資本公積合計')
            WriteToExcel(excelPath, workSheetName,'B', 85, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 85, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 85, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 85, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '保留盈餘':
            WriteToExcel(excelPath, workSheetName,'A', 86, '        保留盈餘')
            WriteToExcel(excelPath, workSheetName,'B', 86, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 86, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 86, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 86, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '法定盈餘公積':
            WriteToExcel(excelPath, workSheetName,'A', 87, '          法定盈餘公積')
            WriteToExcel(excelPath, workSheetName,'B', 87, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 87, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 87, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 87, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '特別盈餘公積':
            WriteToExcel(excelPath, workSheetName,'A', 88, '          特別盈餘公積')
            WriteToExcel(excelPath, workSheetName,'B', 88, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 88, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 88, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 88, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '未提撥保留盈餘':
            WriteToExcel(excelPath, workSheetName,'A', 89, '          未提撥保留盈餘')
            WriteToExcel(excelPath, workSheetName,'B', 89, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 89, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 89, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 89, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '保留盈餘合計':
            WriteToExcel(excelPath, workSheetName,'A', 90, '            保留盈餘合計')
            WriteToExcel(excelPath, workSheetName,'B', 90, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 90, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 90, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 90, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '股東權益其他調整項目合計' and item == 434 :
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 91, '        股東權益其他調整項目合計')
            WriteToExcel(excelPath, workSheetName,'B', 91, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 91, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 91, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 91, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '累積換算調整數':
            WriteToExcel(excelPath, workSheetName,'A', 92, '          累積換算調整數')
            WriteToExcel(excelPath, workSheetName,'B', 92, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 92, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 92, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 92, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '未認列為退休金成本之淨損失':
            WriteToExcel(excelPath, workSheetName,'A', 93, '          未認列為退休金成本之淨損失')
            WriteToExcel(excelPath, workSheetName,'B', 93, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 93, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 93, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 93, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '金融商品之未實現損益':
            WriteToExcel(excelPath, workSheetName,'A', 94, '          金融商品之未實現損益')
            WriteToExcel(excelPath, workSheetName,'B', 94, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 94, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 94, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 94, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '庫藏股票':
            WriteToExcel(excelPath, workSheetName,'A', 95, '          庫藏股票')
            WriteToExcel(excelPath, workSheetName,'B', 95, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 95, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 95, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 95, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '其他股東權益調整項目':
            WriteToExcel(excelPath, workSheetName,'A', 96, '          其他股東權益調整項目')
            WriteToExcel(excelPath, workSheetName,'B', 96, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 96, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 96, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 96, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '股東權益其他調整項目合計' and item == 464:
            #print(item)
            WriteToExcel(excelPath, workSheetName,'A', 97, '            股東權益其他調整項目合計')
            WriteToExcel(excelPath, workSheetName,'B', 97, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 97, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 97, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 97, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '股東權益總計':
            WriteToExcel(excelPath, workSheetName,'A', 98, '          股東權益總計')
            WriteToExcel(excelPath, workSheetName,'B', 98, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 98, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 98, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 98, allTD[item + 4].text.strip())

        if allTD[item].text.strip() == '母公司暨子公司所持有之母公司庫藏股數(單位:股)':
            WriteToExcel(excelPath, workSheetName,'A', 99, '          母公司暨子公司所持有之母公司庫藏股數(單位:股)')
            WriteToExcel(excelPath, workSheetName,'B', 99, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 99, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 99, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 99, allTD[item + 4].text.strip())


        if allTD[item].text.strip() == '預收股款(股東權益項下)之約當發行股數(單位:股)':
            WriteToExcel(excelPath, workSheetName,'A', 100, '          預收股款(股東權益項下)之約當發行股數(單位:股)')
            WriteToExcel(excelPath, workSheetName,'B', 100, allTD[item + 1].text.strip())
            WriteToExcel(excelPath, workSheetName,'C', 100, allTD[item + 2].text.strip())
            WriteToExcel(excelPath, workSheetName,'D', 100, allTD[item + 3].text.strip())
            WriteToExcel(excelPath, workSheetName,'E', 100, allTD[item + 4].text.strip())

for i in range(1000,9999):
    #檢查股票是否存在

    StockIsExist = CheckStockExist(i)

    if StockIsExist == True:
    #產生EXCEL檔案
        filePath = CreateExcelFile("C:\\Users\\user\\Desktop\\python", i)
    # 對excel檔案寫入資料
        DownloadData(i, filePath, i)
        Time.sleep(10)



# # import openpyxl
# #
# # workbook = openpyxl.load_workbook('data.xlsx')
# # sheet = workbook.active